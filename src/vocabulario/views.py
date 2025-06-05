from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.contrib.auth import authenticate, login
from .ia.deepseek import ai_message
from django.http import HttpResponse
from .forms import RegistroForm, LoginForm, PalabraForm
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from .models import Palabra, Tipo
from .tables import PalabraTable
from .filter import PalabraFilter
from django.db import IntegrityError
from django.shortcuts import get_object_or_404
import django_tables2 as tables
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from .models import Palabra


def base_view(request):
    # Recupera el historial de la sesión o crea uno nuevo
    return render(request, 'inicio.html')

def chat_view(request):
    # Esta vista es para manejar la lógica del chat
    mensajes = request.session.get('mensajes', [])

    if request.method == 'POST':
        user_message = request.POST.get('mensaje')
        if user_message:
            # Añade primero el mensaje del usuario
            mensajes.append({"role": "user", "content": user_message})
            # Luego la respuesta de la IA
            respuesta = ai_message(user_message)
            mensajes.append({"role": "assistant", "content": respuesta})
        request.session['mensajes'] = mensajes

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('chat_messages.html', {'mensajes': mensajes}, request=request)
        return JsonResponse({'html': html})
    return render(request, 'chat.html', {'mensajes': mensajes})

def login_view(request):
    form = LoginForm(request.POST or None)
    error = None
    if request.method == 'POST':
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            remember_me = request.POST.get('remember_me')
            try:
                user_obj = User.objects.get(email=email)
                user = authenticate(request, username=user_obj.username, password=password)
                if user is not None:
                    login(request, user)
                    if remember_me:
                        # 30 días (en segundos)
                        request.session.set_expiry(60 * 60 * 24 * 30)
                    else:
                        # Sesión expira al cerrar el navegador
                        request.session.set_expiry(0)
                    return redirect('inicio')
                else:
                    error = "Correo o contraseña incorrectos"
            except User.DoesNotExist:
                error = "Correo o contraseña incorrectos"
    return render(request, 'login.html', {'form': form, 'error': error})

def register_view(request):
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = RegistroForm()
    return render(request, 'register.html', {'form': form})

@login_required
def palabra_view(request):
    error = None
    if request.method == 'POST':
        form = PalabraForm(request.POST)
        if form.is_valid():
            datos = form.cleaned_data
            existe = Palabra.objects.filter(
                palabra=datos['palabra'],
                pinyin=datos['pinyin'],
                tipo=datos['tipo'],
                nivel_hsk=datos.get('nivel_hsk', 1),
                usuario=request.user
            ).exists()
            if existe:
                error = "¡Esa palabra ya existe con esos datos!"
            else:
                try:
                    Palabra.objects.create(
                        palabra=datos['palabra'],
                        pinyin=datos['pinyin'],
                        traduccion=datos['traduccion'],
                        tipo=datos['tipo'],
                        nivel_hsk=datos.get('nivel_hsk', 1),
                        ejemplo=datos.get('ejemplo', ''),
                        usuario=request.user
                    )
                    return redirect('palabra')
                except IntegrityError:
                    error = "¡Esa palabra ya existe!"
    else:
        form = PalabraForm()

    return render(request, 'palabra.html', {
        'form': form,
        'error': error
    })

@login_required
def mostrar_palabras_view(request):
    palabra_filter = PalabraFilter(request.GET, queryset=Palabra.objects.filter(usuario=request.user))
    table = PalabraTable(palabra_filter.qs)
    tables.RequestConfig(request, paginate={"per_page": 10}).configure(table)
    return render(request, 'mostrar_palabras.html', {
        'table': table,
        'user': request.user,
        'filter': PalabraFilter(request.GET, queryset=Palabra.objects.filter(usuario=request.user)),
    })

def eliminar_palabra_view(request, pk):
    palabra = get_object_or_404(Palabra, pk=pk, usuario=request.user)
    palabra.delete()
    return redirect('mostrar_palabras')

def editar_palabra_view(request, pk):
    palabra = get_object_or_404(Palabra, pk=pk, usuario=request.user)
    if request.method == 'POST':
        form = PalabraForm(request.POST, instance=palabra)
        if form.is_valid():
            form.save()
            return redirect('mostrar_palabras')
    else:
        form = PalabraForm(instance=palabra)

    return render(request, 'palabra.html', {
        'form': form,
        'editar': True,
    })

def exportar_palabras_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Palabras"

    headers = ['Carácter', 'Pinyin', 'Traducción', 'Nivel HSK', 'Ejemplo']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    row_num = 2
    palabras = Palabra.objects.filter(usuario=request.user).select_related('tipo').order_by('tipo__nombre', 'palabra')
    tipos = Tipo.objects.all().order_by('nombre')

    for tipo in tipos:
        grupo = palabras.filter(tipo=tipo)
        if grupo.exists():
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
            cell = ws.cell(row=row_num, column=1)
            cell.value = tipo.nombre
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill(start_color=tipo.color.replace('#', ''), end_color=tipo.color.replace('#', ''), fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            row_num += 1

            for palabra in grupo:
                ws.append([
                    palabra.palabra,
                    palabra.pinyin,
                    palabra.traduccion,
                    palabra.nivel_hsk,
                    palabra.ejemplo or ''
                ])
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
                row_num += 1

    # Ajustar ancho de columnas
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=palabras_agrupadas.xlsx'
    wb.save(response)
    return response
